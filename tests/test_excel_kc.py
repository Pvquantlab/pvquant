"""K-C Excel deltası, mühür 1 (v2.186): Daily-Summary + Accuracy-ReportCard.

Kararlar (kullanıcı, bu oturum):
· K-C1: iki mühür — bu dosya birinci yarıyı kilitler.
· K-C2: beslenecek veri yoksa sayfa HİÇ eklenmez (spec'in "—" varsayılanından
  bilinçli sapma).
· K-C3: ghi_daily yerine poa_gunluk_kwh_m2 (ctx'te GHI yok, POA var —
  dürüst adlandırma, uydurma çeviri yok).

Karne satırları GERÇEK üreticiden (_karne_satirlari) gelir; burada semantik
yeniden yazılmaz, yalnız Excel'e düşen yüzey denetlenir (v2.149 dersi:
fikstür elle uydurulmaz). Tarihler bugüne göre kurulur çünkü üretici
'son 30 takvim günü, dünle biter' çapasını now(UTC)'den alır.
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO
from types import SimpleNamespace

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from pvquant.reporting import from_results, build_excel

BUGUN = dt.datetime.now(dt.timezone.utc).date()


# ----------------------------------------------------------------- fikstürler
def _sentetik_forecast(saat=168):
    idx = pd.date_range("2026-07-14 00:00", periods=saat, freq="1h", tz="UTC")
    saat_ici = idx.tz_convert("Europe/Istanbul").hour.values
    gunes = np.clip(np.sin((saat_ici - 5) / 14 * np.pi), 0, None)
    p_ac = gunes * 3800.0
    hourly = pd.DataFrame({
        "poa": gunes * 1050,
        "temp_cell": 25 + 25 * gunes,
        "p_dc_kw": p_ac * 1.03,
        "p_ac_kw": p_ac,
        "energy_kwh": p_ac,
    }, index=idx)
    plant = SimpleNamespace(p_nom_kwp=4514.0, latitude=37.87, longitude=32.49,
                            tilt=30.0, azimuth=180.0)
    meta = {"power_model": "barhdadi_bennis", "meteo_source": "open-meteo"}
    return SimpleNamespace(hourly=hourly, plant=plant, meta=meta)


def _ctx(**ekstra):
    ctx = from_results(_sentetik_forecast(), None, plant_name="Test")
    for k, v in ekstra.items():
        setattr(ctx, k, v)
    return ctx


def _karne_df(gunler_geri, skill_pct=37.7, naif=15.4):
    """Gerçek üretici kolonlarıyla skill_daily fotoğrafı; gün başına 0-24 +
    24-72 kovası. gunler_geri: bugünden geriye gün sayıları listesi."""
    satir = []
    for g in gunler_geri:
        t = pd.Timestamp(BUGUN - dt.timedelta(days=g))
        satir.append({"date": t, "horizon_bucket": "0-24", "mape": 9.6,
                      "naive_wmape": naif, "skill_vs_naive": skill_pct})
        satir.append({"date": t, "horizon_bucket": "24-72", "mape": 13.0,
                      "naive_wmape": None, "skill_vs_naive": None})
    return pd.DataFrame(satir)


def _wb(ctx):
    return load_workbook(BytesIO(build_excel(ctx)), data_only=False)


# ------------------------------------------------------------ Daily-Summary
def test_daily_summary_basliklar_ve_p50():
    ctx = _ctx()
    wb = _wb(ctx)
    assert "Daily-Summary" in wb.sheetnames
    ws = wb["Daily-Summary"]
    basliklar = [c.value for c in ws[1]]
    assert basliklar == ["tarih", "p50_mwh", "p10_mwh", "p90_mwh",
                         "kapasite_faktoru_pct", "hucre_sicaklik_max_c",
                         "poa_gunluk_kwh_m2"]
    # K-C3: ghi hiçbir başlıkta geçmez — dürüst adlandırma
    assert not any("ghi" in str(b).lower() for b in basliklar)
    # 7 gün sözleşmesi + p50 toplamı ctx ile aynı sayı
    assert ws.max_row == 1 + len(ctx.daily_kwh)
    p50_toplam = sum(ws.cell(r, 2).value for r in range(2, ws.max_row + 1))
    assert abs(p50_toplam - ctx.total_mwh) < 0.01


def test_daily_summary_bantsizda_p10_p90_bos():
    ws = _wb(_ctx())["Daily-Summary"]
    for r in range(2, ws.max_row + 1):
        assert ws.cell(r, 3).value is None      # boşluk, "—" metni DEĞİL
        assert ws.cell(r, 4).value is None


def test_daily_summary_bantli_p10_p90_dolu():
    ctx = _ctx()
    ctx.daily_p10 = ctx.daily_kwh * 0.9
    ctx.daily_p90 = ctx.daily_kwh * 1.1
    ws = _wb(ctx)["Daily-Summary"]
    for r in range(2, ws.max_row + 1):
        p50, p10, p90 = (ws.cell(r, c).value for c in (2, 3, 4))
        assert p10 is not None and p90 is not None
        assert p10 < p50 < p90


def test_daily_summary_kf_ve_poa_makul():
    ctx = _ctx()
    ws = _wb(ctx)["Daily-Summary"]
    for r in range(2, ws.max_row + 1):
        kf = ws.cell(r, 5).value
        poa = ws.cell(r, 7).value
        tmax = ws.cell(r, 6).value
        assert kf is not None and 0.0 <= kf <= 100.0
        assert poa is not None and 0.0 <= poa <= 15.0     # kWh/m²·gün
        assert tmax is not None and 20.0 <= tmax <= 60.0


# ---------------------------------------------------- Accuracy-ReportCard
def test_karne_yoksa_sayfa_yok():
    wb = _wb(_ctx())
    assert "Accuracy-ReportCard" not in wb.sheetnames     # K-C2
    # mevcut üç sayfa + Daily-Summary yerli yerinde, sıra bozulmadı
    assert wb.sheetnames == ["Ozet", "Saatlik", "Daily-Summary", "Metadata"]


def test_karne_olculu_gun_yoksa_sayfa_yok():
    # tüm günler 30 günlük pencerenin DIŞInda → üretici ValueError → sayfa yok
    ctx = _ctx(karne=_karne_df([45, 40]))
    wb = _wb(ctx)
    assert "Accuracy-ReportCard" not in wb.sheetnames     # K-C2
    assert "Daily-Summary" in wb.sheetnames               # kalanlar sağlam


def test_karne_sayfasi_30_satir_ve_tek_kaynak_degerleri():
    ctx = _ctx(karne=_karne_df([1, 2, 3]))
    wb = _wb(ctx)
    assert wb.sheetnames == ["Ozet", "Saatlik", "Daily-Summary",
                             "Accuracy-ReportCard", "Metadata"]
    ws = wb["Accuracy-ReportCard"]
    assert [c.value for c in ws[1]] == ["tarih", "wmape_0_24", "wmape_24_72",
                                        "naif_wmape", "skill", "olculdu",
                                        "kapsama_pct"]
    assert ws.max_row == 1 + 30                 # TAM 30 takvim satırı (v2.140)
    dolu = [r for r in range(2, ws.max_row + 1)
            if ws.cell(r, 6).value is True]
    assert len(dolu) == 3
    r = dolu[0]
    assert abs(ws.cell(r, 2).value - 9.6) < 1e-9          # wmape_0_24
    assert abs(ws.cell(r, 3).value - 13.0) < 1e-9         # wmape_24_72
    assert abs(ws.cell(r, 4).value - 15.4) < 1e-9         # naif_wmape
    assert abs(ws.cell(r, 5).value - 0.377) < 1e-3        # skill 0-1 kesir


def test_karne_olculmemis_gun_bos_ve_false():
    ctx = _ctx(karne=_karne_df([1]))
    ws = _wb(ctx)["Accuracy-ReportCard"]
    bos = [r for r in range(2, ws.max_row + 1)
           if ws.cell(r, 6).value is False]
    assert len(bos) == 29
    for r in bos:                               # false ⇒ null (D18 yönü)
        for c in (2, 3, 4, 5):
            assert ws.cell(r, c).value is None
        assert ws.cell(r, 1).value is not None  # tarih hep dolu


def test_karne_kapsama_yayilir_ve_esik_disi_gun_duser():
    from pvquant.config import get_settings
    esik = get_settings().karne_kapsama_esik_pct
    g1 = str(BUGUN - dt.timedelta(days=1))
    g2 = str(BUGUN - dt.timedelta(days=2))
    ctx = _ctx(karne=_karne_df([1, 2]),
               karne_kapsama={g1: 100, g2: esik - 1})
    ws = _wb(ctx)["Accuracy-ReportCard"]
    satir = {str(ws.cell(r, 1).value)[:10]: r
             for r in range(2, ws.max_row + 1)}
    assert ws.cell(satir[g1], 6).value is True
    assert ws.cell(satir[g1], 7).value == 100
    # eşik altı gün: ölçüm arşivde ama karne DIŞI (C-3b) — false + boş
    assert ws.cell(satir[g2], 6).value is False
    assert ws.cell(satir[g2], 2).value is None
    assert ws.cell(satir[g2], 7).value == esik - 1        # kapsama gizlenmez


# ---------------------------------------- Calibration + Climate (v2.187, 2/2)
def test_kalibrasyon_ve_iklim_yoksa_sayfalar_yok():
    wb = _wb(_ctx())
    assert "Calibration" not in wb.sheetnames          # K-C2
    assert "Climate" not in wb.sheetnames              # K-C2


def test_sayfa_sirasi_tam_yedi():
    ctx = _ctx(karne=_karne_df([1]), eta_bos=0.897,
               iklim=pd.DataFrame({"yil": [2024], "ay": [6],
                                   "ghi_kwh_m2": [210.0]}))
    wb = _wb(ctx)
    assert wb.sheetnames == ["Ozet", "Saatlik", "Daily-Summary",
                             "Accuracy-ReportCard", "Calibration",
                             "Climate", "Metadata"]    # spec §4 sırası


def test_kalibrasyon_alanlar_ve_yokluk_tire():
    ctx = _ctx(eta_bos=0.897, bg=0.151, n_valid_hours=1487,
               mape_pct=13.6, coverage_pct=96.4)
    ws = _wb(ctx)["Calibration"]
    assert [c.value for c in ws[1]] == ["alan", "deger", "birim"]
    d = {ws.cell(r, 1).value: ws.cell(r, 2).value
         for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
    assert abs(d["eta_bos"] - 0.897) < 1e-9
    assert abs(d["bifacial_bg"] - 0.151) < 1e-9        # ham katsayı (v2.133)
    assert d["gecerli_saat"] == 1487
    assert d["holdout_mape_pct"] == "\u2014"           # yokluk gizlenmez
    assert d["kalibrasyon_tarihi"] == "\u2014"


def test_kalibrasyon_bayrak_tablosu_ayri():
    ctx = _ctx(eta_bos=0.9,
               flag_dagilimi={"valid": 4100, "night": 2900, "frozen": 12})
    ws = _wb(ctx)["Calibration"]
    hucre = {(r, c): ws.cell(r, c).value
             for r in range(1, ws.max_row + 1) for c in (1, 2)}
    bas = next(r for (r, c), v in hucre.items() if v == "bayrak")
    assert ws.cell(bas, 2).value == "satir_sayisi"
    govde = {ws.cell(bas + i, 1).value: ws.cell(bas + i, 2).value
             for i in range(1, 4)}
    assert govde == {"valid": 4100, "night": 2900, "frozen": 12}
    assert ws.cell(bas + 1, 1).value == "valid"        # çoktan aza sıralı


def test_iklim_beklenti_bicimi():
    ik = pd.DataFrame({"ay": range(1, 13),
                       "ghi_p10_kwh_m2": [80.0] * 12,
                       "ghi_p50_kwh_m2": [110.0] * 12,
                       "ghi_p90_kwh_m2": [140.0] * 12})
    ws = _wb(_ctx(iklim=ik))["Climate"]
    assert [c.value for c in ws[1]][:4] == ["ay", "ghi_p10_kwh_m2",
                                            "ghi_p50_kwh_m2",
                                            "ghi_p90_kwh_m2"]
    assert ws.cell(2, 1).value == 1 and ws.cell(13, 1).value == 12
    assert abs(ws.cell(2, 3).value - 110.0) < 1e-9


def test_iklim_tarihce_bicimi_ham_basilir_yuzdelik_hesaplanmaz():
    ik = pd.DataFrame({"yil": [2024, 2024, 2025],
                       "ay": [6, 7, 6],
                       "ghi_kwh_m2": [205.0, 215.0, 198.0]})
    ws = _wb(_ctx(iklim=ik))["Climate"]
    basliklar = [c.value for c in ws[1] if c.value]
    assert basliklar == ["yil", "ay", "ghi_kwh_m2"]    # ham tarihçe
    assert not any(b and "p50" in str(b) for b in basliklar)  # hesap YOK
    assert ws.max_row >= 4


def test_iklim_yalniz_son12_ile_sayfa_var():
    aylar = pd.date_range("2025-09-01", periods=12, freq="MS")
    s12 = pd.DataFrame({"ay": aylar,
                        "actual_mwh": [float(300 + i) for i in range(12)]})
    ws = _wb(_ctx(son12=s12))["Climate"]
    assert ws.cell(1, 1).value == "ay"
    assert ws.cell(1, 2).value == "gercek_mwh"
    assert abs(ws.cell(2, 2).value - 300.0) < 1e-9
    assert ws.cell(13, 2).value == 311.0
    # actual_ghi kolonu YOK — kaynağı yok, uydurulmadı (K-C3 ilkesi)
    assert not any("ghi" in str(c.value).lower() for c in ws[1] if c.value)


def test_iklim_taninmayan_bicim_tablo_basmaz_son12_kalir():
    ik = pd.DataFrame({"garip_kolon": [1, 2]})
    aylar = pd.date_range("2026-01-01", periods=2, freq="MS")
    s12 = pd.DataFrame({"ay": aylar, "actual_mwh": [310.0, 320.0]})
    wb = _wb(_ctx(iklim=ik, son12=s12))
    ws = wb["Climate"]
    assert ws.cell(1, 1).value == "ay"                 # doğrudan son12 tablosu
    assert ws.cell(1, 2).value == "gercek_mwh"
    # tanınmayan biçim TEK başınaysa sayfa da yok
    wb2 = _wb(_ctx(iklim=pd.DataFrame({"garip_kolon": [1]})))
    assert "Climate" not in wb2.sheetnames
